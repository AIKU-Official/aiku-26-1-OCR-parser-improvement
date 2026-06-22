"""
PDF Structure Suite — PyQt5 GUI
Layout+OCR 모드 (PPStructureV3 전체) + OCR only 모드 (Recognition만)
마크다운 출력 및 미리보기 지원
"""
import sys
import os
import json
import time
import re
import traceback
from pathlib import Path
from datetime import datetime
from typing import List

# ── Paddle 환경변수: import 전에 설정 ────────────────────────────────────────
_MODEL_HOME = Path(os.path.expanduser("~")) / ".paddlex"
os.environ["PADDLE_PDX_CACHE_HOME"]                 = str(_MODEL_HOME)
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"
os.environ["DISABLE_MODEL_SOURCE_CHECK"]            = "True"
os.environ["PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT"]    = "False"
os.environ["CUDA_VISIBLE_DEVICES"]                  = ""

import fitz
import numpy as np
from PIL import Image, ImageDraw
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTextEdit, QProgressBar,
    QFrame, QMessageBox, QStatusBar, QCheckBox, QRadioButton,
    QButtonGroup, QTabWidget
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QDragEnterEvent, QDropEvent


# ─── GPU 감지 ────────────────────────────────────────────────────────────────

def detect_gpu() -> bool:
    try:
        import paddle
        return paddle.device.is_compiled_with_cuda() and paddle.device.cuda.device_count() > 0
    except Exception:
        return False


# ─── 마크다운 생성 ────────────────────────────────────────────────────────────

def _html_table_to_md(html: str) -> str:
    from html.parser import HTMLParser
    class TP(HTMLParser):
        def __init__(self):
            super().__init__(); self.rows=[]; self.row=[]; self.cell=""; self.in_cell=False
        def handle_starttag(self,t,a):
            if t=="tr": self.row=[]
            elif t in("td","th"): self.in_cell=True; self.cell=""
        def handle_endtag(self,t):
            if t in("td","th"): self.row.append(self.cell.strip()); self.in_cell=False
            elif t=="tr":
                if self.row: self.rows.append(self.row)
        def handle_data(self,d):
            if self.in_cell: self.cell+=d
    tp = TP(); tp.feed(html)
    if not tp.rows: return ""
    md = []
    for i, row in enumerate(tp.rows):
        md.append("| " + " | ".join(str(c).replace("|","\\|") for c in row) + " |")
        if i == 0:
            md.append("| " + " | ".join("---" for _ in row) + " |")
    return "\n".join(md)


def build_markdown(pages_data: List[dict], source_pdf: str, mode: str) -> str:
    stem  = Path(source_pdf).stem
    lines = [f"# {stem}\n"]
    lines.append(f"> 처리 방식: **{mode}**  \n")
    lines.append(f"> 처리 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  \n")
    lines.append(f"> 총 페이지: {len(pages_data)}  \n\n---\n")

    for pg in pages_data:
        pnum = pg.get("page_index", 0) + 1
        lines.append(f"\n## 페이지 {pnum}\n")

        full_text = pg.get("full_text", "").strip()
        if full_text:
            lines.append(full_text + "\n")

        for i, tbl in enumerate(pg.get("tables", []), 1):
            lines.append(f"\n### 표 {i}\n")
            md_table = tbl.get("markdown", "") or _html_table_to_md(tbl.get("html",""))
            if md_table:
                lines.append(md_table + "\n")

        for i, fig in enumerate(pg.get("figures", []), 1):
            crop_path = fig.get("crop_path", "")
            caption   = fig.get("caption", f"그림 {i}")
            if crop_path:
                lines.append(f"\n### {caption}\n")
                lines.append(f"![{caption}]({crop_path})\n")

    return "\n".join(lines)


# ─── Layout 파싱 헬퍼 (PPStructureV3 결과 파싱) ───────────────────────────────

_TEXT_LABELS  = frozenset({
    "text","title","paragraph","header","footer","caption","reference",
    "equation","abstract","doc_title","paragraph_title","table_title",
    "table_footnote","figure_title","number","algorithm","content_area",
})
_TABLE_LABELS = frozenset({"table"})
_IMAGE_LABELS = frozenset({"figure","image","chart","figure_caption","seal"})
_HTML_KEYS    = ("block_content","html","table_html","res_html","content")


def _try_json(result):
    try:
        data = result.json
        if callable(data): data = data()
        if isinstance(data, dict): return data
    except Exception:
        pass
    return {}

def _extract_parsing_list(raw_results):
    for result in raw_results:
        data = _try_json(result)
        if data:
            for path in (["parsing_res_list"], ["res","parsing_res_list"]):
                node = data
                for key in path:
                    node = node.get(key) if isinstance(node, dict) else None
                if isinstance(node, list): return node
    return []

def _first_result_json(raw_results):
    for r in raw_results:
        d = _try_json(r)
        if d: return d
    return {}

def _label(block): return str(block.get("block_label", block.get("label", block.get("type","unknown")))).lower()
def _bbox(block):
    for key in ("block_bbox","coordinate","bbox","box","region"):
        val = block.get(key)
        if isinstance(val,(list,tuple)) and len(val)>=4:
            return [float(v) for v in val[:4]]
    return [0.,0.,0.,0.]
def _text(block):
    for key in ("block_content","text","rec_text","content","ocr_text"):
        val = block.get(key)
        if val: return str(val)
    return ""
def _extract_table_html(raw_block):
    for key in _HTML_KEYS:
        val = raw_block.get(key,"")
        if val: return str(val)
    for sub in raw_block.get("sub_blocks",[]):
        for key in _HTML_KEYS:
            val = sub.get(key,"")
            if val: return str(val)
    return ""

def extract_ocr_lines(raw_data):
    ocr    = raw_data.get("overall_ocr_res") or raw_data.get("res",{}).get("overall_ocr_res",{}) or {}
    texts  = ocr.get("rec_texts") or []
    scores = ocr.get("rec_scores") or []
    boxes  = ocr.get("rec_boxes") or []
    lines  = []
    for idx, text in enumerate(texts):
        line = {
            "id":         f"l{idx+1:03d}",
            "text":       str(text),
            "confidence": round(float(scores[idx]),4) if idx < len(scores) else 0.0,
        }
        if idx < len(boxes): line["bbox"] = [float(v) for v in boxes[idx][:4]]
        lines.append(line)
    return lines

def parse_page_result(raw_results, page_index, page_image_path):
    meta = {"width": 0, "height": 0}
    try:
        img  = Image.open(page_image_path)
        w, h = img.size
        meta = {"width": w, "height": h}
    except Exception:
        pass

    raw_data     = _first_result_json(raw_results)
    parsing_list = _extract_parsing_list(raw_results)
    ocr_lines    = extract_ocr_lines(raw_data)

    blocks = []
    table_counter = image_counter = 0
    page_num = page_index + 1

    for raw in parsing_list:
        label         = _label(raw)
        bbox          = _bbox(raw)
        reading_order = len(blocks) + 1
        block_id      = f"p{page_num:03d}_b{reading_order:03d}"

        if label in _TEXT_LABELS:
            blocks.append({"id":block_id,"type":"text","label":label,
                           "text":_text(raw),"bbox":bbox,"reading_order":reading_order})
        elif label in _TABLE_LABELS:
            table_counter += 1
            name = f"page_{page_num:03d}_table_{table_counter:03d}"
            blocks.append({"id":block_id,"type":"table","label":label,"bbox":bbox,
                           "html_path":f"tables/{name}.html",
                           "html":_extract_table_html(raw),
                           "reading_order":reading_order})
        elif label in _IMAGE_LABELS:
            image_counter += 1
            name = f"page_{page_num:03d}_image_{image_counter:03d}"
            blocks.append({"id":block_id,"type":"image","label":label,"bbox":bbox,
                           "crop_path":f"crops/{name}.png",
                           "html_path":f"images/{name}.html",
                           "reading_order":reading_order})
        else:
            blocks.append({"id":block_id,"type":label,
                           "text":_text(raw),"bbox":bbox,"reading_order":reading_order})

    return blocks, meta, ocr_lines, raw_data


# ─── 시각화 / 저장 헬퍼 ──────────────────────────────────────────────────────

_VIS_COLOR = {
    "table":(255,165,0,160),"figure":(34,139,34,160),"image":(34,139,34,160),
    "chart":(34,180,120,160),"title":(70,130,180,160),"text":(30,120,220,120),
}
_TYPE_KO = {
    "table":"표","figure":"그림","image":"이미지","chart":"차트",
    "title":"제목","text":"텍스트","header":"헤더","footer":"푸터",
}

def _draw_blocks(pil_img, blocks):
    draw = ImageDraw.Draw(pil_img, "RGBA")
    for b in blocks:
        bbox  = b.get("bbox",[])
        btype = b.get("type","text")
        color = _VIS_COLOR.get(btype,(120,120,120,120))
        if len(bbox) < 4: continue
        x1,y1,x2,y2 = int(bbox[0]),int(bbox[1]),int(bbox[2]),int(bbox[3])
        draw.rectangle([x1,y1,x2,y2], fill=color)
        draw.rectangle([x1,y1,x2,y2], outline=(*color[:3],255), width=2)
        label = _TYPE_KO.get(btype, btype)
        lw = len(label)*13+8
        draw.rectangle([x1,y1,x1+lw,y1+20], fill=(*color[:3],220))
        draw.text((x1+4,y1+2), label, fill=(255,255,255))
    return pil_img

def _save_table_html(block, base):
    html_rel = block.get("html_path","")
    if not html_rel: return
    body = block.get("html","") or "<p><em>No content</em></p>"
    out  = base/html_rel
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<style>table{{border-collapse:collapse;width:100%}}"
        f"td,th{{border:1px solid #999;padding:6px}}th{{background:#eef4fb}}</style></head>"
        f"<body>{body}</body></html>",
        encoding="utf-8"
    )

def _save_table_xlsx(block, base):
    html_rel = block.get("html_path","")
    html_str = block.get("html","")
    if not html_rel or not html_str: return
    from html.parser import HTMLParser
    class TP(HTMLParser):
        def __init__(self):
            super().__init__(); self.rows=[]; self.row=[]; self.cell=""; self.in_cell=False
        def handle_starttag(self,t,a):
            if t=="tr": self.row=[]
            elif t in("td","th"): self.in_cell=True; self.cell=""
        def handle_endtag(self,t):
            if t in("td","th"): self.row.append(self.cell.strip()); self.in_cell=False
            elif t=="tr":
                if self.row: self.rows.append(self.row)
        def handle_data(self,d):
            if self.in_cell: self.cell+=d
    tp = TP(); tp.feed(html_str)
    if not tp.rows: return
    xlsx_path = base/html_rel.replace(".html",".xlsx")
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = Path(html_rel).stem[:31]
    hf    = Font(bold=True,color="FFFFFF")
    hfill = PatternFill("solid",fgColor="4A90D9")
    thin  = Side(style="thin",color="CCCCCC")
    bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)
    rf    = PatternFill("solid",fgColor="EEF4FB")
    for ri,row in enumerate(tp.rows,1):
        for ci,val in enumerate(row,1):
            c = ws.cell(row=ri,column=ci,value=val); c.border=bdr
            c.alignment=Alignment(vertical="center",wrap_text=True)
            if ri==1: c.font=hf; c.fill=hfill
            elif ri%2==0: c.fill=rf
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=18
    wb.save(str(xlsx_path))

def _save_image_files(block, page_image_path, base):
    crop_rel = block.get("crop_path","")
    html_rel = block.get("html_path","")
    if not crop_rel: return
    bbox      = block.get("bbox",[])
    crop_path = base/crop_rel
    crop_path.parent.mkdir(parents=True, exist_ok=True)
    crop_saved = False
    try:
        img = Image.open(page_image_path).convert("RGB")
        if len(bbox) >= 4:
            w,h = img.size
            x1,y1,x2,y2 = max(0,int(bbox[0])),max(0,int(bbox[1])),min(w,int(bbox[2])),min(h,int(bbox[3]))
            if x2>x1 and y2>y1:
                img.crop((x1,y1,x2,y2)).save(str(crop_path),"PNG")
                crop_saved = True
    except Exception:
        pass
    if html_rel:
        name    = Path(crop_rel).stem
        img_src = f"../crops/{Path(crop_rel).name}"
        img_tag = f'<img src="{img_src}" style="max-width:100%">' if crop_saved else "<p>[Crop unavailable]</p>"
        out = base/html_rel
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(
            f"<!DOCTYPE html><html><head><meta charset='utf-8'></head>"
            f"<body><h2>{name}</h2>{img_tag}</body></html>",
            encoding="utf-8"
        )


# ─── OCR Worker ───────────────────────────────────────────────────────────────

class OCRWorker(QThread):
    progress  = pyqtSignal(int, str)
    page_done = pyqtSignal(int, dict)
    finished  = pyqtSignal(dict)
    error     = pyqtSignal(str)

    def __init__(self, pdf_path, output_dir, use_gpu=False,
                 ocr_mode="layout_ocr", output_opts=None):
        super().__init__()
        self.pdf_path    = pdf_path
        self.output_dir  = output_dir
        self.use_gpu     = use_gpu
        self.ocr_mode    = ocr_mode   # "layout_ocr" or "ocr_only"
        self.output_opts = output_opts or {"json":True,"md":True,"html":True,"xlsx":True}

    def run(self):
        try:
            from paddleocr import PPStructureV3, PaddleOCR
            gpu_label  = "GPU (CUDA)" if self.use_gpu else "CPU"
            device     = "gpu:0"  if self.use_gpu else "cpu"
            is_layout  = self.ocr_mode == "layout_ocr"
            mode_label = "Layout+OCR" if is_layout else "OCR only"

            if self.use_gpu: os.environ["CUDA_VISIBLE_DEVICES"] = "0"
            else:            os.environ["CUDA_VISIBLE_DEVICES"] = "-1"

            self.progress.emit(2, f"{mode_label} 초기화 중... [{gpu_label}]")

            if is_layout:
                pipeline = PPStructureV3(
                    device=device,
                    use_doc_orientation_classify=False,
                    use_doc_unwarping=False,
                    use_textline_orientation=False,
                    lang="korean",
                    text_recognition_model_name="korean_PP-OCRv5_mobile_rec",
                )
            else:
                pipeline = PaddleOCR(
                    use_textline_orientation=True,
                    lang="korean",
                    text_recognition_model_name="korean_PP-OCRv5_mobile_rec",
                    device=device,
                )

            stem = Path(self.pdf_path).stem
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            base = Path(self.output_dir) / f"{stem}_{ts}_{self.ocr_mode}"
            for d in ("pages","tables","crops","images","raw"):
                (base/d).mkdir(parents=True, exist_ok=True)

            self.progress.emit(5, "PDF 렌더링 중...")
            doc   = fitz.open(self.pdf_path)
            total = len(doc)
            mat   = fitz.Matrix(150/72, 150/72)
            img_paths = []
            for i, page in enumerate(doc):
                pix = page.get_pixmap(matrix=mat, alpha=False)
                out = base/"pages"/f"page_{i+1:03d}.png"
                pix.save(str(out))
                img_paths.append(str(out))
            doc.close()

            pages_data = []
            html_paths = []

            for i, img_path in enumerate(img_paths):
                pct = int(10 + (i/total)*80)
                self.progress.emit(pct, f"페이지 {i+1}/{total} {mode_label} 중... [{gpu_label}]")

                if is_layout:
                    # Layout+OCR: 구조 분석 + 텍스트 인식
                    raw    = list(pipeline.predict(input=img_path))
                    blocks, meta, ocr_lines, raw_data = parse_page_result(raw, i, img_path)
                else:
                    # OCR only: Recognition 모델만
                    raw_results = list(pipeline.predict(input=img_path))
                    ocr_lines   = []
                    raw_data    = {}
                    for res in raw_results:
                        try:
                            data = res.json() if callable(res.json) else res.json
                            if not isinstance(data, dict): continue
                            raw_data   = data
                            rec_res    = data.get("res", {})
                            texts      = rec_res.get("rec_texts",  []) or []
                            scores     = rec_res.get("rec_scores", []) or []
                            boxes      = rec_res.get("rec_boxes",  []) or rec_res.get("dt_polys",[]) or []
                            for idx, text in enumerate(texts):
                                if str(text).strip():
                                    ocr_lines.append({
                                        "id":         f"l{idx+1:03d}",
                                        "text":       str(text),
                                        "confidence": round(float(scores[idx]),4) if idx < len(scores) else 0.0,
                                        "bbox":       [float(v) for v in boxes[idx][:4]] if idx < len(boxes) and len(boxes[idx])>=4 else [],
                                    })
                        except Exception:
                            continue
                    pil_img = Image.open(img_path)
                    w, h    = pil_img.size
                    meta    = {"width": w, "height": h}
                    blocks  = [{"id":f"p{i+1:03d}_b{j+1:03d}","type":"text",
                                "text":l["text"],"bbox":l.get("bbox",[]),
                                "reading_order":j+1}
                               for j, l in enumerate(ocr_lines)]

                # raw JSON 저장
                (base/"raw"/f"page_{i+1:03d}_raw.json").write_text(
                    json.dumps(raw_data, ensure_ascii=False, indent=2, default=str),
                    encoding="utf-8"
                )

                # 표/그림 저장
                tables  = []
                figures = []
                for block in blocks:
                    if block["type"] == "table":
                        if self.output_opts.get("html", True):
                            _save_table_html(block, base)
                            html_paths.append(str(base/block["html_path"]))
                        if self.output_opts.get("xlsx", True):
                            _save_table_xlsx(block, base)
                        tables.append({
                            "html":      block.get("html",""),
                            "markdown":  _html_table_to_md(block.get("html","")),
                            "html_path": str(base/block["html_path"]) if self.output_opts.get("html",True) else "",
                        })
                    elif block["type"] == "image":
                        _save_image_files(block, img_path, base)
                        figures.append({
                            "crop_path": str(base/block.get("crop_path","")),
                            "caption":   f"그림 {len(figures)+1}",
                        })

                # 바운딩박스 시각화
                pil_img = Image.open(img_path).convert("RGB")
                _draw_blocks(pil_img.copy(), blocks).save(
                    str(base/"pages"/f"page_{i+1:03d}_bbox.png")
                )

                page_data = {
                    "page_index":  i,
                    "image_path":  f"pages/page_{i+1:03d}.png",
                    "width":       meta["width"],
                    "height":      meta["height"],
                    "full_text":   "\n".join(l["text"] for l in ocr_lines if l["text"].strip()),
                    "ocr_lines":   ocr_lines,
                    "blocks":      [{k:v for k,v in b.items() if not k.startswith("_")} for b in blocks],
                    "tables":      tables,
                    "figures":     figures,
                    "table_count": sum(1 for b in blocks if b["type"]=="table"),
                    "image_count": sum(1 for b in blocks if b["type"]=="image"),
                }

                # 페이지별 JSON 저장
                (base/"pages"/f"page_{i+1:03d}.json").write_text(
                    json.dumps(page_data, ensure_ascii=False, indent=2),
                    encoding="utf-8"
                )
                pages_data.append(page_data)
                self.page_done.emit(i+1, page_data)

            # 전체 result.json
            result_data = {
                "source":       {"file_name": Path(self.pdf_path).name, "page_count": total},
                "mode":         self.ocr_mode,
                "processed_at": datetime.now().isoformat(),
                "accelerator":  gpu_label,
                "summary": {
                    "page_count":     total,
                    "table_count":    sum(p["table_count"] for p in pages_data),
                    "image_count":    sum(p["image_count"] for p in pages_data),
                    "ocr_line_count": sum(len(p["ocr_lines"]) for p in pages_data),
                },
                "full_text": "\n\n".join(p["full_text"] for p in pages_data),
                "pages":     pages_data,
            }

            json_path = None
            if self.output_opts.get("json", True):
                self.progress.emit(93, "result.json 저장 중...")
                json_path = base/"result.json"
                json_path.write_text(json.dumps(result_data, ensure_ascii=False, indent=2), encoding="utf-8")
                json_path = str(json_path)

            md_content = build_markdown(pages_data, self.pdf_path, f"{mode_label} ({gpu_label})")
            md_path = None
            if self.output_opts.get("md", True):
                self.progress.emit(96, "마크다운 저장 중...")
                md_path = base/f"{stem}_result.md"
                md_path.write_text(md_content, encoding="utf-8")
                md_path = str(md_path)

            self.progress.emit(100, "완료!")
            self.finished.emit({
                "base_dir":   str(base),
                "json_path":  json_path,
                "md_path":    md_path,
                "html_paths": html_paths,
                "page_count": total,
                "gpu_used":   self.use_gpu,
                "pages":      pages_data,
                "md_content": md_content,
            })

        except Exception:
            self.error.emit(traceback.format_exc())


# ─── Drop Zone ────────────────────────────────────────────────────────────────

class DropZone(QLabel):
    clicked = pyqtSignal()
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("drop_zone")
        self.setAlignment(Qt.AlignCenter)
        self.setText("📄\n\nPDF 파일을 여기에 드래그하거나\n클릭하여 선택하세요")
        self.setFont(QFont("Malgun Gothic", 12))
        self.setAcceptDrops(True)
        self.setCursor(Qt.PointingHandCursor)
        self._pdf_path = None

    def mousePressEvent(self, e): self.clicked.emit()

    def dragEnterEvent(self, e: QDragEnterEvent):
        if e.mimeData().hasUrls():
            if e.mimeData().urls()[0].toLocalFile().lower().endswith(".pdf"):
                e.acceptProposedAction()

    def dropEvent(self, e: QDropEvent):
        path = e.mimeData().urls()[0].toLocalFile()
        if path.lower().endswith(".pdf"): self.set_file(path)

    def set_file(self, path):
        self._pdf_path = path
        name    = Path(path).name
        size_mb = os.path.getsize(path)/1024/1024
        self.setText(f"✅  {name}\n\n{size_mb:.2f} MB")
        self.setStyleSheet("border:2px solid #4A90D9;background:#EEF4FB;color:#1A4A7A;")

    @property
    def pdf_path(self): return self._pdf_path


# ─── Stylesheet ───────────────────────────────────────────────────────────────

STYLE = """
QMainWindow,QWidget{background:#F5F7FA;color:#1E2A38;font-family:'Malgun Gothic','Apple SD Gothic Neo',sans-serif;}
#card{background:#FFFFFF;border:1px solid #D8E3EE;border-radius:12px;padding:8px;}
#drop_zone{background:#EEF4FB;border:2px dashed #94BDE0;border-radius:16px;min-height:130px;color:#5A7A9A;}
#drop_zone:hover{border-color:#4A90D9;background:#E4EFF9;color:#2C5F8A;}
#mode_box{background:#FFFFFF;border:1px solid #D8E3EE;border-radius:10px;padding:10px 16px;}
QPushButton#primary{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #3A80CC,stop:1 #5AA0E8);
  color:white;border:none;border-radius:8px;padding:10px 28px;font-size:14px;font-weight:bold;}
QPushButton#primary:hover{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #2A70BC,stop:1 #4A90D8);}
QPushButton#primary:disabled{background:#BDD0E4;color:#FFFFFF;}
QPushButton#secondary{background:#FFFFFF;color:#4A7AA8;border:1px solid #B8CEE4;
  border-radius:8px;padding:8px 20px;font-size:13px;}
QPushButton#secondary:hover{border-color:#4A90D9;color:#2C5F8A;background:#EEF4FB;}
QPushButton#secondary:disabled{color:#A0B8CC;border-color:#D8E3EE;}
QProgressBar{background:#DDE8F2;border:none;border-radius:6px;height:10px;color:transparent;}
QProgressBar::chunk{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #3A80CC,stop:1 #5ABCEE);border-radius:6px;}
QTextEdit{background:#FAFCFF;border:1px solid #D0DFF0;border-radius:8px;
  color:#2A4A6A;font-family:'Consolas','D2Coding',monospace;font-size:12px;padding:6px;}
QTabWidget::pane{border:1px solid #D0DFF0;border-radius:8px;background:#FAFCFF;}
QTabBar::tab{background:#EEF4FB;border:1px solid #D0DFF0;padding:6px 16px;
  border-top-left-radius:6px;border-top-right-radius:6px;color:#4A7AA8;}
QTabBar::tab:selected{background:#FFFFFF;color:#1E5A9A;font-weight:bold;border-bottom:none;}
QScrollBar:vertical{background:#EEF4FB;width:8px;border-radius:4px;}
QScrollBar::handle:vertical{background:#A8C8E8;border-radius:4px;min-height:20px;}
QStatusBar{background:#E8EFF7;color:#6A8AA8;font-size:11px;border-top:1px solid #C8D8E8;}
QCheckBox{color:#3A5A7A;font-size:12px;spacing:6px;}
QCheckBox::indicator{width:16px;height:16px;border-radius:4px;border:1.5px solid #94BDE0;background:#FFFFFF;}
QCheckBox::indicator:checked{background:#3A80CC;border-color:#3A80CC;}
QCheckBox:disabled{color:#A0B8CC;}
QRadioButton{color:#3A5A7A;font-size:13px;spacing:6px;}
QRadioButton::indicator{width:16px;height:16px;border-radius:8px;border:1.5px solid #94BDE0;background:#FFFFFF;}
QRadioButton::indicator:checked{background:#3A80CC;border-color:#3A80CC;}
QLabel#title{font-size:21px;font-weight:bold;color:#1E5A9A;letter-spacing:1px;}
QLabel#subtitle{font-size:11px;color:#7A9ABB;letter-spacing:0.5px;}
QLabel#stat_val{font-size:18px;font-weight:bold;color:#2A6AAE;}
QLabel#stat_lbl{font-size:10px;color:#8AAABB;}
QLabel#gpu_on{background:#E0F2E9;color:#1A7A48;border:1px solid #7ECBA0;border-radius:6px;padding:3px 10px;font-size:11px;font-weight:bold;}
QLabel#gpu_off{background:#F0F0F0;color:#778899;border:1px solid #CCCCCC;border-radius:6px;padding:3px 10px;font-size:11px;}
"""


# ─── Main Window ──────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PDF Structure Suite")
        self.setMinimumSize(980, 780)
        self.worker         = None
        self._output_dir    = str(Path.home()/"Documents"/"PDFStructureResults")
        self._gpu_available = detect_gpu()
        self._last_base_dir = None
        self._build_ui()
        self.setStyleSheet(STYLE)

    def _build_ui(self):
        central = QWidget(); self.setCentralWidget(central)
        root = QVBoxLayout(central); root.setContentsMargins(24,20,24,12); root.setSpacing(12)

        # ── 헤더 ──
        hdr = QHBoxLayout(); tc = QVBoxLayout()
        t = QLabel("PDF STRUCTURE SUITE"); t.setObjectName("title")
        s = QLabel("Layout+OCR (구조 분석 + 텍스트 인식)  /  OCR only (텍스트 인식만)  ·  JSON + Markdown + XLSX")
        s.setObjectName("subtitle")
        tc.addWidget(t); tc.addWidget(s); hdr.addLayout(tc); hdr.addStretch()

        badge = QLabel("🟢  GPU 감지됨" if self._gpu_available else "⚪  CPU 전용")
        badge.setObjectName("gpu_on" if self._gpu_available else "gpu_off")
        hdr.addWidget(badge); hdr.addSpacing(12)

        self.stat_pages   = self._stat_widget("0","페이지")
        self.stat_tables  = self._stat_widget("0","표")
        self.stat_figures = self._stat_widget("0","그림")
        self.stat_time    = self._stat_widget("--","소요 시간")
        for s in [self.stat_pages,self.stat_tables,self.stat_figures,self.stat_time]:
            hdr.addWidget(s)
        root.addLayout(hdr)

        # ── OCR 모드 선택 ──
        mode_box = QFrame(); mode_box.setObjectName("mode_box")
        mode_lay = QHBoxLayout(mode_box); mode_lay.setContentsMargins(12,8,12,8)
        mode_lbl = QLabel("OCR 모드:"); mode_lbl.setStyleSheet("font-weight:bold;color:#1E5A9A;")
        self.radio_layout = QRadioButton("Layout + OCR  (구조 분석 + 텍스트 인식)")
        self.radio_ocr    = QRadioButton("OCR only  (텍스트 인식만 — 빠름)")
        self.radio_layout.setChecked(True)
        self.btn_group = QButtonGroup()
        self.btn_group.addButton(self.radio_layout)
        self.btn_group.addButton(self.radio_ocr)
        self.chk_gpu = QCheckBox("GPU 가속")
        self.chk_gpu.setChecked(self._gpu_available)
        self.chk_gpu.setEnabled(self._gpu_available)
        mode_lay.addWidget(mode_lbl)
        mode_lay.addWidget(self.radio_layout)
        mode_lay.addWidget(self.radio_ocr)
        mode_lay.addSpacing(20)
        mode_lay.addWidget(self.chk_gpu)
        mode_lay.addStretch()
        root.addWidget(mode_box)

        # ── 출력 형식 선택 ──
        out_opt_box = QFrame(); out_opt_box.setObjectName("mode_box")
        out_opt_lay = QHBoxLayout(out_opt_box); out_opt_lay.setContentsMargins(12,8,12,8)
        opt_lbl = QLabel("출력 형식:"); opt_lbl.setStyleSheet("font-weight:bold;color:#1E5A9A;")
        self.chk_out_json = QCheckBox("JSON")
        self.chk_out_md   = QCheckBox("Markdown")
        self.chk_out_html = QCheckBox("HTML (표)")
        self.chk_out_xlsx = QCheckBox("Excel (표)")
        for chk in [self.chk_out_json, self.chk_out_md, self.chk_out_html, self.chk_out_xlsx]:
            chk.setChecked(True)
        out_opt_lay.addWidget(opt_lbl)
        out_opt_lay.addWidget(self.chk_out_json)
        out_opt_lay.addWidget(self.chk_out_md)
        out_opt_lay.addWidget(self.chk_out_html)
        out_opt_lay.addWidget(self.chk_out_xlsx)
        out_opt_lay.addStretch()
        root.addWidget(out_opt_box)

        # ── 드롭존 ──
        self.drop_zone = DropZone(); self.drop_zone.clicked.connect(self._pick_pdf)
        root.addWidget(self.drop_zone)

        # ── 저장 경로 ──
        out_row = QHBoxLayout()
        self.lbl_out = QLabel(f"저장 폴더:  {self._output_dir}")
        self.lbl_out.setStyleSheet("color:#7A9ABB;font-size:11px;"); self.lbl_out.setWordWrap(True)
        btn_out = QPushButton("📁 변경"); btn_out.setObjectName("secondary")
        btn_out.setFixedWidth(80); btn_out.clicked.connect(self._pick_output)
        out_row.addWidget(self.lbl_out,1); out_row.addWidget(btn_out)
        root.addLayout(out_row)

        # ── 진행바 ──
        self.progress_bar = QProgressBar(); self.progress_bar.setValue(0); self.progress_bar.setFixedHeight(10)
        root.addWidget(self.progress_bar)
        self.lbl_status = QLabel(""); self.lbl_status.setStyleSheet("color:#7A9ABB;font-size:11px;")
        root.addWidget(self.lbl_status)

        # ── 탭 ──
        self.tabs = QTabWidget()
        log_widget = QWidget(); log_lay = QVBoxLayout(log_widget); log_lay.setContentsMargins(6,6,6,6)
        self.log = QTextEdit(); self.log.setReadOnly(True); self.log.setMinimumHeight(180)
        log_lay.addWidget(self.log)
        self.tabs.addTab(log_widget, "📋 로그")

        md_widget = QWidget(); md_lay = QVBoxLayout(md_widget); md_lay.setContentsMargins(6,6,6,6)
        self.md_preview = QTextEdit(); self.md_preview.setReadOnly(True)
        self.md_preview.setPlaceholderText("처리 완료 후 마크다운 결과가 여기에 표시됩니다.")
        self.md_preview.setMinimumHeight(180)
        md_lay.addWidget(self.md_preview)
        self.tabs.addTab(md_widget, "📝 마크다운 미리보기")
        root.addWidget(self.tabs, 1)

        # ── 버튼 ──
        btn_row = QHBoxLayout()
        self.btn_run     = QPushButton("▶  분석 시작");      self.btn_run.setObjectName("primary")
        self.btn_open    = QPushButton("📂 결과 폴더 열기"); self.btn_open.setObjectName("secondary")
        self.btn_copy_md = QPushButton("📋 마크다운 복사");  self.btn_copy_md.setObjectName("secondary")
        self.btn_run.clicked.connect(self._run)
        self.btn_open.clicked.connect(self._open_output);    self.btn_open.setEnabled(False)
        self.btn_copy_md.clicked.connect(self._copy_md);     self.btn_copy_md.setEnabled(False)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_copy_md)
        btn_row.addWidget(self.btn_open)
        btn_row.addWidget(self.btn_run)
        root.addLayout(btn_row)

        self.status_bar = QStatusBar(); self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("PDF 파일을 선택하고 모드를 선택한 후 분석을 시작하세요.")

    def _stat_widget(self,val,label):
        w=QFrame(); w.setObjectName("card")
        lay=QVBoxLayout(w); lay.setContentsMargins(14,8,14,8); lay.setSpacing(2)
        v=QLabel(val); v.setObjectName("stat_val"); v.setAlignment(Qt.AlignCenter)
        l=QLabel(label); l.setObjectName("stat_lbl"); l.setAlignment(Qt.AlignCenter)
        lay.addWidget(v); lay.addWidget(l); w._val_label=v; return w

    def _pick_pdf(self):
        path,_=QFileDialog.getOpenFileName(self,"PDF 선택","","PDF Files (*.pdf)")
        if path: self.drop_zone.set_file(path); self.status_bar.showMessage(f"선택됨: {path}")

    def _pick_output(self):
        folder=QFileDialog.getExistingDirectory(self,"저장 폴더",self._output_dir)
        if folder: self._output_dir=folder; self.lbl_out.setText(f"저장 폴더:  {folder}")

    def _log(self,msg):
        ts=datetime.now().strftime("%H:%M:%S")
        self.log.append(f"<span style='color:#6AAAD8;'>[{ts}]</span>  {msg}")

    def _copy_md(self):
        QApplication.clipboard().setText(self.md_preview.toPlainText())
        self.status_bar.showMessage("마크다운이 클립보드에 복사됐어요!")

    def _run(self):
        if not self.drop_zone.pdf_path:
            QMessageBox.warning(self,"파일 없음","PDF를 먼저 선택해주세요."); return

        use_layout = self.radio_layout.isChecked()
        use_gpu    = self.chk_gpu.isChecked() and self._gpu_available
        ocr_mode   = "layout_ocr" if use_layout else "ocr_only"
        output_opts = {
            "json": self.chk_out_json.isChecked(),
            "md":   self.chk_out_md.isChecked(),
            "html": self.chk_out_html.isChecked(),
            "xlsx": self.chk_out_xlsx.isChecked(),
        }

        self.btn_run.setEnabled(False); self.btn_open.setEnabled(False); self.btn_copy_md.setEnabled(False)
        self.progress_bar.setValue(0); self.log.clear(); self.md_preview.clear()
        self._start_time=time.time(); self._total_tables=0; self._total_figs=0
        for s in [self.stat_pages,self.stat_tables,self.stat_figures]: s._val_label.setText("0")
        self.stat_time._val_label.setText("--")

        mode  = "Layout+OCR" if use_layout else "OCR only"
        accel = "GPU" if use_gpu else "CPU"
        self._log(f"시작: <b style='color:#1A4A8A;'>{self.drop_zone.pdf_path}</b>")
        self._log(f"모드: <b style='color:#1E5A9A;'>{mode}</b>  |  가속기: <b>{accel}</b>")
        self._log(f"저장: {self._output_dir}")

        self.worker = OCRWorker(
            self.drop_zone.pdf_path, self._output_dir,
            use_gpu=use_gpu, ocr_mode=ocr_mode, output_opts=output_opts
        )
        self.worker.progress.connect(self._on_progress)
        self.worker.page_done.connect(self._on_page_done)
        self.worker.finished.connect(self._on_finished)
        self.worker.error.connect(self._on_error)
        self.worker.start()
        self.tabs.setCurrentIndex(0)

    def _on_progress(self,pct,msg):
        self.progress_bar.setValue(pct); self.lbl_status.setText(msg)
        self.stat_time._val_label.setText(f"{time.time()-self._start_time:.1f}s")

    def _on_page_done(self,page_num,data):
        self._total_tables += data.get("table_count",0)
        self._total_figs   += data.get("image_count",0)
        self.stat_pages._val_label.setText(str(page_num))
        self.stat_tables._val_label.setText(str(self._total_tables))
        self.stat_figures._val_label.setText(str(self._total_figs))
        self._log(f"페이지 {page_num} — 표 {data.get('table_count',0)}개  그림 {data.get('image_count',0)}개  텍스트 {len(data.get('ocr_lines',[]))}줄")

    def _on_finished(self,result):
        elapsed=time.time()-self._start_time
        self.stat_time._val_label.setText(f"{elapsed:.1f}s")
        self._last_base_dir=result["base_dir"]
        self._log(f"<b style='color:#1A7A48;'>✅ 완료!</b>  {elapsed:.1f}초")
        if result.get("json_path"): self._log(f"JSON     : {result['json_path']}")
        if result.get("md_path"):   self._log(f"Markdown : {result['md_path']}")
        self._log(f"저장 위치 : {result['base_dir']}")
        md = result.get("md_content","")
        if md:
            self.md_preview.setPlainText(md)
            self.tabs.setCurrentIndex(1)
            self.btn_copy_md.setEnabled(True)
        self.lbl_status.setText("완료!")
        self.btn_run.setEnabled(True); self.btn_open.setEnabled(True)
        self.status_bar.showMessage(f"완료  ·  {result['base_dir']}")

    def _on_error(self,tb):
        self._log(f"<span style='color:#CC3333;'>오류:</span><br><pre style='color:#AA2222;'>{tb}</pre>")
        self.btn_run.setEnabled(True); self.lbl_status.setText("오류 — 로그 확인")
        self.status_bar.showMessage("오류 발생")
        self.tabs.setCurrentIndex(0)

    def _open_output(self):
        if self._last_base_dir and os.path.exists(self._last_base_dir):
            if sys.platform=="win32": os.startfile(self._last_base_dir)
            else: os.system(f"open '{self._last_base_dir}'")


def main():
    app=QApplication(sys.argv); app.setStyle("Fusion")
    win=MainWindow(); win.show(); sys.exit(app.exec_())

if __name__=="__main__":
    main()
